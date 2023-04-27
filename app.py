from flask import Flask, request, send_file, render_template
import pandas as pd
import openai
import openpyxl

openai.api_key = "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"

app = Flask(__name__)


@app.route('/')
def index():
    return render_template('index.html')


# Load the ChatGPT model
model = "text-davinci-002"
prompt = "Generate a mapping of elements to activities in retail fitouts."


# Define a function to generate the prompt
def generate_prompt(element_names, element_descs, element_cats, element_sections, element_item_types, activity_names,
                    activity_cats):
    prompt_template = "An BOQ element in the construction of retail fitouts contains the name and a description that specifies how to install the element & other requirements. FOr the following element :-\n"
    # prompt_template += f"Element Description: {element_descs}\n"
    prompt_template += f"Element {1}:\n- Name: {element_names}\n- Description: {element_descs}\n"
    # prompt_template += "List of activities:\n"
    # n = 0
    # # for i in range(10, 20):
    # #     prompt_template += f"- Activity {n + 1}: {activity_names[i]}, Category: {activity_cats[i]}\n"
    # #     n = n + 1
    # # for i in range(55, 60):
    # #     prompt_template += f"- Activity {n + 1}: {activity_names[i]}, Category: {activity_cats[i]}\n"
    # #     n = n + 1
    # for i in range(60, 80):
    #     prompt_template += f"- Activity {n + 1}: {activity_names[i]}, Category: {activity_cats[i]}\n"
    #     n = n + 1
    # prompt_template += f"- Activity {i + 1}: {activity_names[i]}\n"
    prompt_template += "\nPredict the standard activity name to which this elements belongs to and some keywords :."
    prompt_template += "\nwhat category of work in retail fitouts does this elements description says also provide set of keywords for this:."
    return prompt_template


def addColumn(element_wb):
    new_column_index = element_wb.max_column + 1
    header = 'New Column Header'
    element_wb.cell(row=1, column=new_column_index, value=header)


# Define a function to generate the output mapping
def generate_mapping(prompt, tasks_file, index):
    # Generate the response from OpenAI's API
    response = openai.Completion.create(
        engine=model,
        prompt=prompt,
        max_tokens=2097,
        n=1,
        stop=None,
        temperature=0.5,
    )
    # Extract the generated mapping from the response
    mapping = response.choices[0].text.strip()
    element_wb = openpyxl.load_workbook(tasks_file)
    addColumn(element_wb)

    for row in element_wb.active.iter_rows(min_row=2, values_only=True):
        row.append(mapping)
    # Return the mapping as a string
    element_wb.save()

    return mapping


# 'element_name': row[2],
#             'description': row[3],
#             'element_category': row[5],
#             'section': row[1],
#             'item_type': row[6]
# Define a function to load Excel files and extract the necessary data
def load_data(element_file, activity_file):
    # Load the Excel files using openpyxl
    element_wb = openpyxl.load_workbook(element_file)
    activity_wb = openpyxl.load_workbook(activity_file)
    # Extract the necessary data from the Excel files
    element_names = []
    element_descs = []
    element_cats = []
    element_sections = []
    element_item_types = []
    for row in element_wb.active.iter_rows(min_row=2, values_only=True):
        element_names.append(row[2])
        element_descs.append(row[3])
        element_cats.append(row[5])
        element_sections.append(row[1])
        element_item_types.append(row[6])
    activity_names = []
    activity_cats = []
    for row in activity_wb.active.iter_rows(min_row=2, values_only=True):
        activity_names.append(row[0])
        activity_cats.append(row[1])
    # Return the extracted data as a tuple
    return (
        element_names, element_descs, element_cats, element_sections, element_item_types, activity_names, activity_cats)


# Define the Flask app and endpoints
from flask import Flask, request, jsonify

app = Flask(__name__)


@app.route('/generate_mapping', methods=['POST'])
def generate_mapping_endpoint():
    # Load the input files and extract the necessary data
    if 'tasks_file' and 'activities_file' not in request.files:
        return 'No tasks_file and activities_file file provided', 400

    file = request.files['tasks_file']
    if file.filename == '':
        return 'No file provided', 400

    if not file.filename.endswith('.xlsx'):
        return 'Invalid file format', 400

    # Load input Excel files
    tasks_file = request.files['tasks_file']
    activities_file = request.files['activities_file']
    data = load_data(tasks_file, activities_file)
    for i in range(0, 100):
        prompt = generate_prompt(data[0][i], data[1][i], data[2][i], data[3][i], data[4][i], data[5], data[6])
        mapping = generate_mapping(prompt, tasks_file, i)


# @app.route('/predict-activity', methods=['POST'])
# def predict_activity():
#     if 'file' not in request.files:
#         return 'No file provided', 400
#
#     file = request.files['file']
#     if file.filename == '':
#         return 'No file provided', 400
#
#     if not file.filename.endswith('.xlsx'):
#         return 'Invalid file format', 400
#
#     tasks_df = pd.read_excel(file, sheet_name='tasks')
#     activities_df = pd.read_excel(file, sheet_name='activities')
#
#     # Generate prompts for each task
#     prompts = []
#     for index, row in tasks_df.iterrows():
#         prompt = f"Predict which activity '{row['name']}' belongs to:\n\n"
#         for activity in activities_df['activity']:
#             prompt += f"{activity}\n"
#         prompts.append(prompt)
#
#     # Generate predictions for each task
#     predictions = []
#     for prompt in prompts:
#         response = openai.Completion.create(
#             engine="text-davinci-002",
#             prompt=prompt,
#             max_tokens=512,
#             n=1,
#             stop=None,
#             temperature=0.5,
#         )
#         prediction = response.choices[0].text.strip()
#         predictions.append(prediction)
#
#     # Add predictions to tasks dataframe
#     tasks_df['activity'] = predictions
#
#     # Save predicted tasks to Excel file
#     predicted_file = openpyxl.Workbook()
#     predicted_sheet = predicted_file.active
#
#     for r in dataframe_to_rows(tasks_df, index=False, header=True):
#         predicted_sheet.append(r)
#
#     predicted_file.save('predicted_tasks.xlsx')
#
#     # Return Excel file
#     return send_file('predicted_tasks.xlsx',
#                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#
#
# def dataframe_to_rows(df, index=False, header=True):
#     if header:
#         yield df.columns.tolist()
#     for _, row in df.iterrows():
#         yield row.tolist()
model = "text-davinci-002"
prompt = "Generate a mapping of elements to activities in retail fitouts."

# # Define endpoint for handling POST requests
# @app.route('/predict', methods=['POST'])
# def predict():
#     if 'tasks_file' and 'activities_file' not in request.files:
#         return 'No tasks_file and activities_file file provided', 400
#
#     file = request.files['tasks_file']
#     if file.filename == '':
#         return 'No file provided', 400
#
#     if not file.filename.endswith('.xlsx'):
#         return 'Invalid file format', 400
#
#     # Load input Excel files
#     tasks_file = request.files['tasks_file']
#     activities_file = request.files['activities_file']
#     tasks_workbook = openpyxl.load_workbook(tasks_file)
#     activities_workbook = openpyxl.load_workbook(activities_file)
#     tasks_sheet = tasks_workbook.active
#     activities_sheet = activities_workbook.active
#
#     # Extract data from input Excel files
#     tasks = []
#     for row in tasks_sheet.iter_rows(min_row=2, values_only=True):
#         tasks.append({
#             'element_name': row[2],
#             'description': row[3],
#             'element_category': row[5],
#             'section': row[1],
#             'item_type': row[6]
#         })
#
#     activities = []
#     for row in activities_sheet.iter_rows(min_row=2, values_only=True):
#         activities.append({
#             'activity': row[0],
#             'category': row[1]
#         })
#
#     # Generate prompts for each task and activity combination
#     prompt_responses = []
#     for task in tasks:
#         for activity in activities:
#             prompt_template = "An element in the construction of retail fitouts contains the name and a description that specifies how to install the element, and also has a category. These elements are grouped in several activities to be performed, each activity can contain one or more elements based on the element's 'name', 'description', 'category', 'section', and 'item type'.\n\nPlease generate a mapping of elements to activities based on the following information:\n\n"
#             # prompt = f"Which activity does the task '{task['element_name']}' belong to? Activity: '{activity['activity']}' Category: '{activity['category']}'"
#             response = openai.Completion.create(engine="davinci", prompt=prompt, max_tokens=1024)
#             prompt_responses.append({
#                 'task_element_name': task['element_name'],
#                 'task_description': task['description'],
#                 'task_element_category': task['element_category'],
#                 'task_section': task['section'],
#                 'task_item_type': task['item_type'],
#                 'activity': activity['activity'],
#                 'category': activity['category'],
#                 'response': response.choices[0].text.strip()
#             })
#
#     # Generate output Excel file containing mapping of tasks to activities
#     output_workbook = openpyxl.Workbook()
#     output_sheet = output_workbook.active
#     output_sheet.append([
#         'Task Element Name',
#         'Task Description',
#         'Task Element Category',
#         'Task Section',
#         'Task Item Type',
#         'Activity',
#         'Category',
#         'Prediction'
#     ])
#     for prompt_response in prompt_responses:
#         output_sheet.append([
#             prompt_response['task_element_name'],
#             prompt_response['task_description'],
#             prompt_response['task_element_category'],
#             prompt_response['task_section'],
#             prompt_response['task_item_type'],
#             prompt_response['activity'],
#             prompt_response['category'],
#             prompt_response['response']
#         ])
#
#     output_file = 'output.xlsx'
#     output_workbook.save(output_file)
#
#     return f'Output saved to {output_file}'


if __name__ == '__main__':
    app.run(debug=True)
